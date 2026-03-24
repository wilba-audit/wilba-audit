"""Generate WILBA Subscription Costs Dashboard as Excel."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Subscriptions"

# Styles
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
category_font = Font(name="Calibri", bold=True, size=11, color="1A1A1A")
category_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
currency_format = '$#,##0.00'
total_font = Font(name="Calibri", bold=True, size=12)
total_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
thin_border = Border(
    bottom=Side(style="thin", color="DDDDDD")
)
thick_border = Border(
    top=Side(style="medium", color="2D2D2D"),
    bottom=Side(style="medium", color="2D2D2D"),
)

# Column widths
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 35

# Title
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value = "WILBA — Subscription & Tool Costs Dashboard"
title_cell.font = Font(name="Calibri", bold=True, size=16, color="1A1A1A")
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 35

ws.merge_cells("A2:F2")
ws["A2"].value = "Fill in your actual monthly costs. Yellow cells = needs your input."
ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="888888")

# Headers
headers = ["Service", "Category", "Monthly Cost (USD)", "Currency", "Annual Cost", "Notes"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[4].height = 28

# Highlight fill for cells that need input
input_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

# Subscription data: (service, category, estimated_cost, currency, notes)
subscriptions = [
    # AI & API Services
    ("CATEGORY", "AI & API Services", "", "", ""),
    ("Anthropic (Claude API)", "AI & API", None, "USD", "Powers Claude Code, Telegram bot, audit pipeline"),
    ("OpenAI", "AI & API", None, "USD", "Voice transcription for Telegram bot"),
    ("ElevenLabs", "AI & API", None, "USD", "Voice cloning — Content Generation Machine"),
    ("HeyGen", "AI & API", None, "USD", "Avatar video — Content Generation Machine"),
    ("Gemini (Google)", "AI & API", None, "USD", "Used via API key in .env"),
    ("Deepgram", "AI & API", None, "USD", "Speech-to-text API"),
    ("Perplexity", "AI & API", None, "USD", "Research — Content Generation Machine"),

    # Content & Social
    ("CATEGORY", "Content & Social Media", "", "", ""),
    ("CreatorMate", "Content", None, "USD", "Final video output — Content Machine"),
    ("Descript", "Content", None, "USD", "Podcast clipping for Danielle retainer"),
    ("Canva", "Content", None, "USD", "Design — check if you have Pro"),

    # Business & CRM
    ("CATEGORY", "Business & CRM", "", "", ""),
    ("GoHighLevel", "CRM", None, "USD", "CRM — check if active or paused"),
    ("Kajabi", "CRM", None, "USD", "Danielle's email + course platform"),
    ("Notion", "Productivity", None, "USD", "Workspace — may be free tier"),
    ("Airtable", "Productivity", None, "USD", "May be free tier"),

    # Hosting & Infrastructure
    ("CATEGORY", "Hosting & Infrastructure", "", "", ""),
    ("Render", "Hosting", 0, "USD", "Audit pipeline — free tier currently?"),
    ("Cloudflare", "DNS", 0, "USD", "DNS for wilba.ai — free tier"),
    ("Porkbun", "Domain", 10, "USD", "wilba.ai domain — ~$10/year = $0.83/mo"),
    ("Wix", "Website", None, "USD", "wilba.ai website hosting"),

    # Communication
    ("CATEGORY", "Communication", "", "", ""),
    ("Slack", "Comms", None, "USD", "Multiple workspaces (IHF, main)"),
    ("Telegram", "Comms", 0, "USD", "Free — bot hosting is via your Mac"),

    # Email & Marketing
    ("CATEGORY", "Email & Marketing", "", "", ""),
    ("Resend", "Email", None, "USD", "Transactional email — audit pipeline"),
    ("Kit (ConvertKit)", "Email", None, "USD", "Email marketing — API key in .env"),
    ("Bitly", "Marketing", None, "USD", "Link shortening"),

    # Data & Automation
    ("CATEGORY", "Data & Automation", "", "", ""),
    ("Apify", "Automation", None, "USD", "Web scraping/automation"),
    ("Fireflies.ai", "Meetings", None, "USD", "Meeting transcription"),
    ("Calendly", "Scheduling", None, "USD", "Booking link — may be free tier"),
    ("Google Workspace", "Productivity", None, "USD", "Gmail, Sheets, Drive"),
    ("Supadata", "Data", None, "USD", "API in .env — check if active"),

    # Learning & Community
    ("CATEGORY", "Learning & Community", "", "", ""),
    ("Skool (1% AI / AAA)", "Learning", None, "USD", "AAA Accelerator community"),
    ("YouTube Premium", "Content", None, "USD", "If applicable"),

    # Hardware & Other
    ("CATEGORY", "Other / Hardware Subscriptions", "", "", ""),
    ("Adobe Creative Cloud", "Software", None, "USD", "If applicable — Lightroom/Premiere?"),
    ("iCloud", "Storage", None, "USD", "Apple storage plan"),
    ("Spotify / Music", "Personal", None, "USD", "If applicable"),
]

row = 5
data_start_row = 5
cost_cells = []

for item in subscriptions:
    service, category, cost, currency, notes = item

    if service == "CATEGORY":
        # Category header row
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = category_fill
            cell.font = category_font
        ws.cell(row=row, column=1, value=category)
        ws.row_dimensions[row].height = 24
        row += 1
        continue

    ws.cell(row=row, column=1, value=service)
    ws.cell(row=row, column=2, value=category)

    cost_cell = ws.cell(row=row, column=3)
    if cost is not None:
        cost_cell.value = cost
    else:
        cost_cell.value = 0
        cost_cell.fill = input_fill  # Highlight = needs input
    cost_cell.number_format = currency_format

    ws.cell(row=row, column=4, value=currency if currency else "USD")

    # Annual cost formula
    annual_cell = ws.cell(row=row, column=5)
    annual_cell.value = f"=C{row}*12"
    annual_cell.number_format = currency_format

    ws.cell(row=row, column=6, value=notes)

    # Light border
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border

    cost_cells.append(row)
    row += 1

# Totals row
row += 1
for col in range(1, 7):
    cell = ws.cell(row=row, column=col)
    cell.fill = total_fill
    cell.font = total_font
    cell.border = thick_border

ws.cell(row=row, column=1, value="TOTAL MONTHLY")
ws.cell(row=row, column=2, value="All Services")

# Sum only the cost rows (skip category headers)
cost_refs = ",".join([f"C{r}" for r in cost_cells])
total_monthly = ws.cell(row=row, column=3)
total_monthly.value = f"=SUM({cost_refs})"
total_monthly.number_format = currency_format

annual_refs = ",".join([f"E{r}" for r in cost_cells])
total_annual = ws.cell(row=row, column=5)
total_annual.value = f"=SUM({annual_refs})"
total_annual.number_format = currency_format

ws.cell(row=row, column=6, value="Update yellow cells with your actual costs")

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 18

ws2["A1"].value = "WILBA Cost Summary"
ws2["A1"].font = Font(name="Calibri", bold=True, size=14)

summary_items = [
    ("Total Monthly Spend", f"=Subscriptions!C{row}"),
    ("Total Annual Spend", f"=Subscriptions!E{row}"),
    ("", ""),
    ("Revenue (Current)", ""),
    ("Danielle Retainer", 1500),
    ("Monkey Joe's (avg/mo)", 1033),
    ("WILBA Direct", 0),
    ("Total Monthly Revenue", "=SUM(B7:B9)"),
    ("", ""),
    ("Net Monthly (Revenue - Costs)", "=B10-B3"),
]

for i, (label, value) in enumerate(summary_items, 3):
    ws2.cell(row=i, column=1, value=label).font = Font(name="Calibri", bold=bool(label and "Total" in label or "Net" in label), size=11)
    cell = ws2.cell(row=i, column=2, value=value)
    if isinstance(value, (int, float)) or (isinstance(value, str) and value.startswith("=")):
        cell.number_format = currency_format

output_path = "/Users/jessmorrell/Documents/AAA/aios-starter-kit/outputs/wilba-subscription-costs.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
